import openpyxl
import datetime


# Log class
class Log:

    # Initiializes the class
    def __init__(self, name="log"):
        self.path = name + ".xlsx"
        self.open()
    
    # Opens the excel file if it already exists
    # otherwise a new excel file is created
    def open(self):
        try:
            self.wb = openpyxl.load_workbook(self.path)
            self.sheet = self.wb.active
        except:
            self.create()
    
    # Creates excel file
    def create(self):
        self.wb = openpyxl.Workbook()
        self.wb['Sheet'].title = self.path.split('.')[0]

        self.sheet = self.wb.active

        self.sheet.cell(row=1, column=1).value = "Date"
        self.sheet.cell(row=1, column=2).value = "Time"
        self.sheet.cell(row=1, column=3).value = "Message"

    # Writes message with the current date and time
    #  to a new line in the excel file
    def write(self, message):
        row = len(list(self.sheet.rows)) + 1
        self.sheet.cell(row=row, column=1).value = datetime.datetime.now().strftime("%a - %d %b %Y")
        self.sheet.cell(row=row, column=2).value = datetime.datetime.now().strftime("%H:%M:%S")
        for i in range(len(message)):
            self.sheet.cell(row=row, column=i+3).value = message[i]

    # Returns cell value at row and column
    def read(self, row, column):
        return self.sheet.cell(row=row, column=column).value

    # Saves file
    # If this fails the program asks to retry
    def save(self):
        try:
            self.wb.save(self.path)
        except:
            print("Error saving file")
            resave = input("Do you want to try again? (Yes/No)\n")
            if (resave in ["Yes", "yes", "y"]):
                self.save()



if __name__ == "__main__":
    print("running log")

    log = Log()
    log.write("Dummy data 1")
    log.write("Dummy data 2")
    log.save()