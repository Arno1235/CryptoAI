# Discord and Excel libraries
from discord import Webhook, RequestsWebhookAdapter
import openpyxl
import datetime



# Sends a message to the user over on Discord
def notifyUser(message):
    webhook = Webhook.from_url(getDiscordURL(), adapter=RequestsWebhookAdapter())
    webhook.send(message)

# Returns the discord url from the secret.txt file
def getDiscordURL():
    with open ("secret.txt", "r") as myfile:
        data = myfile.read().splitlines()
        return data[2]

# Log class
class Log:

    # Initiializes the class
    def __init__(self):
        self.path = "log.xlsx"
        self.open()
    
    # Opens the excel file if it already exists
    # otherwise a new excel file is created
    def open(self):
        try:
            self.wb = openpyxl.load_workbook(self.path)
            self.sheet = self.wb.active
        except:
            self.create()
    
    # Creates excel file
    def create(self):
        self.wb = openpyxl.Workbook()
        self.wb['Sheet'].title = self.path.split('.')[0]

        self.sheet = self.wb.active

        self.sheet.cell(row=1, column=1).value = "Date"
        self.sheet.cell(row=1, column=2).value = "Time"
        self.sheet.cell(row=1, column=3).value = "Message"

    # Writes message with the current dat and time
    #  to a new line in the excel file
    def write(self, message):
        row = len(list(self.sheet.rows)) + 1
        self.sheet.cell(row=row, column=1).value = datetime.datetime.now().strftime("%a - %d %b %Y")
        self.sheet.cell(row=row, column=2).value = datetime.datetime.now().strftime("%H:%M:%S")
        self.sheet.cell(row=row, column=3).value = message

    # Returns cell value at row and column
    def read(self, row, column):
        return self.sheet.cell(row=row, column=column).value

    # Saves file
    # If this fails the program asks to retry
    def save(self):
        try:
            self.wb.save(self.path)
        except:
            print("Error saving file")
            resave = input("Do you want to try again? (Yes/No)\n")
            if (resave in ["Yes", "yes", "y"]):
                self.save()

if __name__ == "__main__":
    print("running notify_user")

    notifyUser("Test")

    log = Log()
    log.write("Dummy data 1")
    log.write("Dummy data 2")
    log.save()